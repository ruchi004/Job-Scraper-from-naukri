import re
import time
import requests
import openpyxl
from openpyxl import Workbook
from bs4 import BeautifulSoup
from openpyxl import Workbook
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager

job_domain = "python"
loaction = "bangalore"
url = 'https://www.naukri.com/%s-jobs-in-%s?k=%s&l=%s' % (job_domain , loaction , job_domain , loaction)

wb = Workbook()
sheet =  wb.active
sheet.title = "JOBS"
col = 4
l = ['Company' , 'JobTitle' , 'Salary' , 'Skills Required']
for i in range(1 , col+1):
    c = sheet.cell(row = 1, column = i)
    c.value = l[i-1]
wb.save("C:\\Users\\Ruchi\\Desktop\\demo.xlsx")

driver = webdriver.Chrome(ChromeDriverManager().install())
driver.get(url)

job_title = []
comp_title = []
sals = []
skls = []
tipu = []

for nex in range(2,7):
	time.sleep(3)
	src = driver.page_source
	soup = BeautifulSoup(src , 'html.parser')
	results = soup.find('section' , class_ = 'listContainer fleft')

	time.sleep(3)
	comp_name = results.find_all('a' , class_ = 'subTitle ellipsis fleft')
	j_title = results.find_all('a' , class_ = 'title fw500 ellipsis')
	salaries = results.find_all('span' , class_ = 'ellipsis fleft fs12 lh16')

	for i in range(len(comp_name)):
	    comp_title.append(comp_name[i].text)
	    job_title.append(j_title[i].text)
	    #print(comp_name[i].text , '\n')

	for sal in salaries:
	    sals.append(sal.text)

	time.sleep(3)
	skills = results.find_all('ul' , class_ = 'tags has-description')

	for index  , skill in enumerate(skills):
	    nn = skill.find_all('li' , class_ = 'fleft fs12 grey-text lh16 dot')
	    for jj in nn:
	        skls.append(jj.text)
	    tipu.append(str(skls))
	    skls = []

	wb.save("C:\\Users\\Ruchi\\Desktop\\demo.xlsx")
	driver.find_element_by_link_text(str(nex)).click()
	time.sleep(0.5)

row = len(comp_title)
col = 5
print(row)
print(len(job_title) , len(comp_title) , len(sals) , len(tipu))

sheet =  wb.active

k = 1
for i in range(2 , row+2):
	c1 = sheet.cell(row = i , column = 1)
	c1.value = comp_title[i-2]

	c1 = sheet.cell(row = i , column = 2)
	c1.value = job_title[i-2]

	c1 = sheet.cell(row = i , column = 3)
	c1.value = sals[k]
	k+=3

	c1 = sheet.cell(row = i , column = 4)
	c1.value = tipu[i-2]
wb.save("C:\\Users\\Ruchi\\Desktop\\demo.xlsx")
